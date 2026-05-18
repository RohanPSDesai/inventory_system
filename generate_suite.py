import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_msme_inventory_diagnostic():
    # Initialize Master Workbook Container
    wb = openpyxl.Workbook()
    
    # Define Institutional Typography & Colors
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="555555")
    section_font = Font(name=font_family, size=13, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=10, italic=True, color="555555")
    
    # Palette Fills
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")     # Corporate Navy
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")      # Pastel Input Yellow
    calc_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")       # Light Grey Formula Mute
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # Soft Blue Highlight Card
    card_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Olive Card Header
    
    # Borders & Alignments
    thin_side = Side(style='thin', color='BFBFBF')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # =========================================================================
    # TAB 1: MACRO HEALTH CALCULATORS
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Macro Health Calculators"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1['B2'] = "Inventory Health Macro Diagnostic Engine"
    ws1['B2'].font = title_font
    ws1['B3'] = "Instructions: Input your numbers into the Yellow cells. Use the Currency Selector to toggle display modes."
    ws1['B3'].font = subtitle_font
    
    # Currency Selector Control Block
    ws1['B5'] = "Currency Selector Toggle:"
    ws1['B5'].font = bold_font
    ws1['C5'] = "INR"
    ws1['C5'].font = bold_font
    ws1['C5'].fill = input_fill
    ws1['C5'].alignment = align_center
    
    ws1['D5'] = "Internal Exchange Multiplier:"
    ws1['D5'].font = italic_font
    ws1['E5'] = "=IF(C5=\"INR\", 83.50, 1.00)"
    ws1['E5'].font = bold_font
    ws1['E5'].fill = calc_fill
    ws1['E5'].alignment = align_center
    
    ws1['F5'] = "AGGREGATED COMPANY HEALTH SCORE:"
    ws1['F5'].font = bold_font
    ws1['G5'] = "=(F13+F25+F37)/3"
    ws1['G5'].font = Font(name=font_family, size=12, bold=True, color="1B365D")
    ws1['G5'].fill = summary_fill
    ws1['G5'].alignment = align_center
    ws1['G5'].number_format = '0%'
    
    headers_tab1 = ["Data Metric Label Sourcing Field", "Business Meaning / Context Concept", "Client Value Input", "Dynamic Currency Output Value", "Calculated Metric Result", "Calculated Score & Analysis"]
    
    # --- BLOCK 1: ASSET UTILIZATION ---
    ws1['B8'] = "1. ASSET UTILIZATION (INVENTORY TURNOVER RATIO)"
    ws1['B8'].font = section_font
    for col_idx, text in enumerate(headers_tab1, start=2):
        cell = ws1.cell(row=9, column=col_idx, value=text)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = align_center
    ws1.row_dimensions[9].height = 26
    
    calc1_data = [
        ["Annual Cost of Goods Sold (COGS)", "Total manufacturing or buy-in cost of inventory sold during the year (from P&L statement).", 450000, "=C10*$E$5", "-", "Benchmark target: 4.0x - 6.0x"],
        ["Beginning Inventory Value", "Value of warehouse stock sitting on hand on Day 1 of the fiscal year cycle.", 130000, "=C11*$E$5", "-", "Measures starting locked capital."],
        ["Ending Inventory Value", "Value of warehouse stock remaining on hand on the final day of the fiscal cycle.", 120000, "=C12*$E$5", "-", "Measures trailing capital retention."],
        ["Calculated Average Inventory", "Automated baseline representing capital locked in stock over the year: (Start + End) / 2.", "-", "=AVERAGE(D11:D12)", "=D10/D13", "=MIN(1, IF(E13>=4, 1, E13/4))"]
    ]
    for idx, r_data in enumerate(calc1_data):
        r = 10 + idx
        for c, val in enumerate(r_data, start=2):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = regular_font if idx != 3 else bold_font
            cell.border = thin_border
            if c == 2: cell.alignment = align_left
            elif c == 3:
                cell.alignment = align_right
                if type(val) in [int, float]: cell.number_format = '#,##0'; cell.fill = input_fill
            elif c == 4:
                cell.alignment = align_right; cell.fill = calc_fill
                cell.number_format = '#,##0'
            elif c == 5:
                cell.alignment = align_center; cell.fill = calc_fill
                if idx == 3: cell.number_format = '0.0" x"'
            elif c == 6:
                cell.alignment = align_center
                if idx == 3: cell.fill = summary_fill; cell.number_format = '0%'
                else: cell.font = italic_font
        ws1.row_dimensions[r].height = 22

    # --- BLOCK 2: SUPPLY CHAIN RELIABILITY ---
    ws1['B17'] = "2. SUPPLY CHAIN RELIABILITY (STOCKOUT RATE)"
    ws1['B17'].font = section_font
    for col_idx, text in enumerate(headers_tab1, start=2):
        cell = ws1.cell(row=18, column=col_idx, value=text)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = align_center
    ws1.row_dimensions[18].height = 26
    
    calc2_data = [
        ["Total Unfulfilled Orders due to Stockouts", "Count of customer transactions that failed because warehouse racks were empty.", 14, "=C19", "-", "Benchmark target: Under 2.0%"],
        ["Total Orders Placed by Customers", "Absolute volume of all purchase entries received over the timeline.", 200, "=C20", "-", "Establishes statistical size scaling."],
        ["Days with Stockouts on Core SKUs", "Total days where top revenue-generating lines were completely empty.", 12, "=C21", "-", "Highlights continuity vulnerability."],
        ["Total Operational Business Days", "Total scheduled open days across the audited timeframe block.", 250, "=C22", "-", "Normalizes the timeline density."],
        ["Calculated Order Stockout Rate", "Percentage of customer sales orders failed due to missing stock: (Failed / Total).", "-", "-", "=D19/D20", "=MAX(0, IF(E23<=0.02, 1, 1-((E23-0.02)/0.1)))"],
        ["Calculated SKU Rack Emptiness Time", "Timeline availability friction metric: (Stockout Days / Total Business Days).", "-", "-", "=D21/D22", "Indicates structural shelf health."]
    ]
    for idx, r_data in enumerate(calc2_data):
        r = 19 + idx
        for c, val in enumerate(r_data, start=2):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = regular_font if idx not in [4, 5] else bold_font
            cell.border = thin_border
            if c == 2: cell.alignment = align_left
            elif c == 3:
                cell.alignment = align_right
                if type(val) in [int, float]: cell.number_format = '#,##0'; cell.fill = input_fill
            elif c == 4:
                cell.alignment = align_center if idx in [4,5] else align_right; cell.fill = calc_fill
                if idx not in [4, 5]: cell.number_format = '#,##0'
            elif c == 5:
                cell.alignment = align_center; cell.fill = calc_fill
                if idx in [4, 5]: cell.number_format = '0.0%'
            elif c == 6:
                cell.alignment = align_center
                if idx == 4: cell.fill = summary_fill; cell.number_format = '0%'
                else: cell.font = italic_font
        ws1.row_dimensions[r].height = 22

    # --- BLOCK 3: INVENTORY SECURITY (SHRINKAGE) ---
    ws1['B28'] = "3. INVENTORY SECURITY & TRACKING (SHRINKAGE RATE)"
    ws1['B28'].font = section_font
    for col_idx, text in enumerate(headers_tab1, start=2):
        cell = ws1.cell(row=29, column=col_idx, value=text)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = align_center
    ws1.row_dimensions[29].height = 26
    
    calc3_data = [
        ["Book Inventory Financial Value", "Total valuation of stock that SHOULD theoretically exist on paper via billing records.", 153500, "=C30*$E$5", "-", "Benchmark target: Under 1.0%"],
        ["Physical Inventory Count Value", "Real financial value found sitting on shelves during a manual warehouse count.", 150000, "=C31*$E$5", "-", "The factual baseline reality."],
        ["Known Damaged/Spoiled Value", "Cost value of goods written off cleanly due to accidents, spoilage, or expiration logs.", 500, "=C32*$E$5", "-", "Identified operational loss."],
        ["Calculated Gross Discrepancy", "Total tracking deficit gap: Book Value - Physical Value.", "-", "=D30-D31", "-", "Unexplained variance floor."],
        ["Calculated Net Unknown Shrinkage", "Toxic leakage segment representing unaccounted capital loss (theft, unrecorded scrap).", "-", "=D33-D32", "-", "Pure systematic leakage indicator."],
        ["Calculated System Shrinkage Rate", "Core operational index determining audit tracking safety: (Net Loss / Book Value).", "-", "-", "=D34/D30", "=MAX(0, IF(E35<=0.01, 1, 1-((E35-0.01)/0.05)))"]
    ]
    for idx, r_data in enumerate(calc3_data):
        r = 30 + idx
        for c, val in enumerate(r_data, start=2):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = regular_font if idx != 5 else bold_font
            cell.border = thin_border
            if c == 2: cell.alignment = align_left
            elif c == 3:
                cell.alignment = align_right
                if type(val) in [int, float]: cell.number_format = '#,##0'; cell.fill = input_fill
            elif c == 4:
                cell.alignment = align_right; cell.fill = calc_fill
                if idx != 5: cell.number_format = '#,##0'
            elif c == 5:
                cell.alignment = align_center; cell.fill = calc_fill
                if idx == 5: cell.number_format = '0.0%'
            elif c == 6:
                cell.alignment = align_center
                if idx == 5: cell.fill = summary_fill; cell.number_format = '0%'
                else: cell.font = italic_font
        ws1.row_dimensions[r].height = 22

    # Column Formatting for Tab 1
    ws1.column_dimensions['B'].width = 38
    ws1.column_dimensions['C'].width = 44
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 26
    ws1.column_dimensions['F'].width = 24
    ws1.column_dimensions['G'].width = 34

    # =========================================================================
    # TAB 2: OPERATIONAL MICRO VIEW (CONCEPT & ALIGNMENT PLANNER)
    # =========================================================================
    ws2 = wb.create_sheet(title="Micro Concept Planner")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['B2'] = "STEP-BY-STEP INVENTORY CHECKLIST FOR MSME OWNERS"
    ws2['B2'].font = title_font
    ws2['B3'] = "Instructions: Read the scenarios. Select your current practice level in the Yellow cell to auto-generate scores and your action plan."
    ws2['B3'].font = subtitle_font
    
    ws2['B5'] = "YOUR BUSINESS SCORE:"
    ws2['B5'].font = bold_font
    ws2['C5'] = "=SUM(F11, F16, F21, F26, F31, F36)/(6*3)"
    ws2['C5'].font = Font(name=font_family, size=14, bold=True, color="1B365D")
    ws2['C5'].fill = summary_fill
    ws2['C5'].alignment = align_center
    ws2['C5'].number_format = '0%'
    
    headers_tab2 = ["Core Business Area", "How it Works & Why It Saves You Money", "Choose What Best Describes Your Current Business", "Points", "Your Immediate Custom Action Plan"]
    for col_idx, text in enumerate(headers_tab2, start=2):
        cell = ws2.cell(row=8, column=col_idx, value=text)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = align_center
    ws2.row_dimensions[8].height = 26
    
    modules_data = [
        {
            "row": 10,
            "area": "1. Stock Grouping\n(ABC Analysis)",
            "why": "Splitting your stock into 3 buckets so you know where your cash is locked up:\n• 'A' Items: Your top 20% products that bring in 80% of your profit. Need tight, daily control.\n• 'B' Items: Medium priority.\n• 'C' Items: Cheap bulk items (screws, packing tape) where running out hurts, but tracking every single piece is a waste of time.",
            "default": "1 - Tracked on Excel Spreadsheets",
            "formula": '=IF(LEFT(D11,1)="0", "🔴 CRITICAL ACTION: Spend 1 hour looking at last month\'s sales receipts. List out your top 20% profit-making items. Move them to the front rows of your shop or warehouse so they never go missing.", IF(LEFT(D11,1)="1", "🟡 IMPROVEMENT STEP: Spreadsheets get outdated quickly. Pick your absolute top 5 profit-making items and start tracking them using an automated system alert instead of manual typing.", "🟢 EXCELLENT: Keep reviewing your top earners every 6 months to see if client preferences or items have changed."))'
        },
        {
            "row": 15,
            "area": "2. Restocking Signals\n(Reorder Points)",
            "why": "Knowing the exact moment to order more stock from a supplier so you never have to say 'Out of Stock' to a walk-in client.\n\nFormula is simple: (Items Sold Daily x Days Vendor Takes to Deliver) + a small safety cushion.",
            "default": "0 - Checked by Look-and-Guess",
            "formula": '=IF(LEFT(D16,1)="0", "🔴 CRITICAL ACTION: Avoid guessing by sight. For your top 5 bestsellers, write down exactly how many days it takes your supplier to deliver. Set that volume as your minimum trigger level.", IF(LEFT(D16,1)="1", "🟡 IMPROVEMENT STEP: Build automated alerts into your billing desk. When items drop to your calculated trigger mark, an order sheet should auto-draft.", "🟢 EXCELLENT: Your restocking patterns match real consumer speeds. Audit delivery times with your vendor every quarter."))'
        },
        {
            "row": 20,
            "area": "3. Emergency Cushion\n(Safety Stock)",
            "why": "An intentional extra batch of stock kept strictly hidden away as a backup buffer. This protects your revenue when a supplier delivers late, transport strikes occur, or a sudden festive rush happens.",
            "default": "1 - Tracked on Excel Spreadsheets",
            "formula": '=IF(LEFT(D21,1)="0", "🔴 CRITICAL ACTION: You are running highly exposed to sudden stockouts. Take 2 weeks worth of average sales volume for your best-selling item and physically separate it as emergency reserve.", IF(LEFT(D21,1)="1", "🟡 IMPROVEMENT STEP: Base your backup cushion on real numbers instead of gut feel. Calculate seasonal sales fluctuations to size the buffer efficiently.", "🟢 EXCELLENT: Backup stock is optimized. Only adjust this buffer if your supplier\'s on-time shipping reliability drops below 95%."))'
        },
        {
            "row": 25,
            "area": "4. Stock Rotation\n(FIFO Protocol)",
            "why": "First-In, First-Out rule. Ensuring old stock is moved to the front and sold first before brand new stock packages are opened.\n\nThis keeps items from getting dusty, damaged, expiring, or becoming unsellable dead weight on shelves.",
            "default": "2 - System-Driven Routines",
            "formula": '=IF(LEFT(D26,1)="0", "🔴 CRITICAL ACTION: Rearrange your floor shelves immediately. Make sure your team stacks new inventory from the back so the oldest boxes are naturally picked first by loaders.", IF(LEFT(D26,1)="1", "🟡 IMPROVEMENT STEP: Run random spot checks on your racks once a week. Check batch codes or manufacturing dates to ensure nothing is gathering dust at the bottom.", "🟢 EXCELLENT: Perfect physical stock layout. Your active rotation routine preserves raw capital and avoids dead stock waste."))'
        },
        {
            "row": 30,
            "area": "5. Zero-Storage Flow\n(JIT Feasibility)",
            "why": "Ordering raw goods or products so they arrive *exactly* when an order is ready, keeping your warehouse storage costs near zero.\n\nWarning: This requires extremely fast, 100% reliable local vendors. If they run late, your production line instantly breaks down.",
            "default": "0 - Checked by Look-and-Guess",
            "formula": '=IF(LEFT(D31,1)="0", "🔵 MSME NOTE: Zero-storage is a high-risk approach. Stick to keeping reliable safety buffers unless you have a highly automated local vendor nearby.", IF(LEFT(D31,1)="1", "🔵 MSME NOTE: Test this approach only with your most trustworthy vendor for non-critical packaging items before scaling up.", "🟢 EXCELLENT: Highly optimized local supply chain infrastructure. Keep direct communication lines open to monitor raw vendor timelines daily."))'
        },
        {
            "row": 35,
            "area": "6. Quick Morning Audits\n(Cycle Counting)",
            "why": "Counting just 5 random bins or items every single morning before business doors open.\n\nThis takes 10 minutes and completely replaces the nightmare of shutting down your entire operation for 2 full days at the end of the year to do a massive annual count.",
            "default": "0 - Checked by Look-and-Guess",
            "formula": '=IF(LEFT(D36,1)="0", "🔴 CRITICAL ACTION: Instruct your warehouse lead or store clerk to physically count just 5 random item slots every single morning before opening up. Fix matching errors in your records instantly.", IF(LEFT(D36,1)="1", "🟡 IMPROVEMENT STEP: Create a simple rolling list to ensure every single SKU type gets physically touched and counted at least twice a year.", "🟢 EXCELLENT: Superb operational discipline. Your digital billing system accurately reflects real life on the shelves."))'
        }
    ]
    
    for mod in modules_data:
        r = mod["row"]
        
        # Build clean outer bounding boxes for the card
        for ro in range(4):
            for co in range(2, 7):
                ws2.cell(row=r+ro, column=co).border = thin_border
                
        # Fill Main Category Card Label Row
        c_label = ws2.cell(row=r, column=2, value=mod["area"])
        c_label.font = bold_font; c_label.fill = card_header_fill; c_label.alignment = align_center
        
        # Merge Definition text space cleanly
        ws2.merge_cells(start_row=r, start_column=3, end_row=r+3, end_column=3)
        c_desc = ws2.cell(row=r, column=3, value=mod["why"])
        c_desc.font = regular_font; c_desc.alignment = align_wrap
        
        # Interactive Option Anchor Cell Location
        drop_c = ws2.cell(row=r, column=4, value=mod["default"])
        drop_c.font = bold_font; drop_c.fill = input_fill; drop_c.alignment = align_center
        
        # Dynamic Point Evaluation Multiplier
        score_c = ws2.cell(row=r, column=5, value=f"=IF(LEFT(D{r},1)=\"0\", 0, IF(LEFT(D{r},1)=\"1\", 1, 3))")
        score_c.font = bold_font; score_c.fill = calc_fill; score_c.alignment = align_center
        score_c.number_format = '0'
        
        # Merge and populate Action Text fields
        ws2.merge_cells(start_row=r, start_column=6, end_row=r+3, end_column=6)
        act_c = ws2.cell(row=r, column=6, value=mod["formula"])
        act_c.font = regular_font; act_c.fill = calc_fill; act_c.alignment = align_wrap
        
        # Sub-options Guide Mapping
        ws2.cell(row=r+1, column=4, value="0 - Checked by Look-and-Guess").font = italic_font
        ws2.cell(row=r+1, column=4).alignment = align_left
        ws2.cell(row=r+1, column=5, value="0 Points").font = italic_font
        ws2.cell(row=r+1, column=5).alignment = align_center
        
        ws2.cell(row=r+2, column=4, value="1 - Tracked on Excel Spreadsheets").font = italic_font
        ws2.cell(row=r+2, column=4).alignment = align_left
        ws2.cell(row=r+2, column=5, value="1 Point").font = italic_font
        ws2.cell(row=r+2, column=5).alignment = align_center
        
        ws2.cell(row=r+3, column=4, value="2 - System-Driven Routines").font = italic_font
        ws2.cell(row=r+3, column=4).alignment = align_left
        ws2.cell(row=r+3, column=5, value="3 Points").font = italic_font
        ws2.cell(row=r+3, column=5).alignment = align_center
        
        ws2.row_dimensions[r].height = 28
        ws2.row_dimensions[r+1].height = 22
        ws2.row_dimensions[r+2].height = 22
        ws2.row_dimensions[r+3].height = 22
        
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 54
    ws2.column_dimensions['D'].width = 38
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 68

    # Save finalized workspace container
    file_output = "MSME_Inventory_Consulting_Suite.xlsx"
    wb.save(file_output)
    return file_output

if __name__ == "__main__":
    build_msme_inventory_diagnostic()